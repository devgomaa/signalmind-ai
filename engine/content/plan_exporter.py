"""
engine/content/plan_exporter.py
=================================
Sprint 7: تصدير الـ Content Plan كـ PDF وExcel.
"""

import os
import json
from datetime import datetime
from pathlib import Path

from engine.utils.logger import get_logger

logger = get_logger("PlanExporter")

EXPORT_DIR = "output_exports"


class PlanExporter:
    """
    يصدّر الـ content strategy + trends كـ:
    - PDF  : تقرير احترافي
    - Excel: spreadsheet منظّم بـ sheets
    """

    def __init__(self):
        os.makedirs(EXPORT_DIR, exist_ok=True)

    # ══════════════════════════════════════════
    # PDF Export
    # ══════════════════════════════════════════

    def export_pdf(self, workspace_name: str, niche: str,
                   strategy: str, trends: dict,
                   content_ideas: list = None) -> str:
        """
        يولّد PDF احترافي يحتوي:
        - غلاف بـ workspace name وتاريخ
        - ملخص الترندات (exploding + growing)
        - استراتيجية المحتوى الكاملة
        - Content Ideas لو موجودة
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.colors import HexColor, white, black
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer,
                Table, TableStyle, PageBreak, HRFlowable
            )
            from reportlab.lib.enums import TA_LEFT, TA_CENTER

            return self._export_pdf_reportlab(
                workspace_name, niche, strategy, trends,
                content_ideas or []
            )
        except ImportError:
            logger.warning("reportlab not installed — using plain text PDF fallback")
            return self._export_pdf_plain(workspace_name, niche, strategy, trends)

    def _export_pdf_reportlab(self, ws_name, niche, strategy, trends, ideas) -> str:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer,
            Table, TableStyle, PageBreak, HRFlowable
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER

        timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename   = f"content_plan_{timestamp}.pdf"
        path       = os.path.join(EXPORT_DIR, filename)
        ACCENT     = HexColor("#3B82F6")
        DARK       = HexColor("#0D1320")

        doc   = SimpleDocTemplate(path, pagesize=A4,
                                  leftMargin=2*cm, rightMargin=2*cm,
                                  topMargin=2*cm, bottomMargin=2*cm)
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle("title",
            parent=styles["Heading1"], fontSize=28, spaceAfter=6,
            textColor=ACCENT, alignment=TA_CENTER)
        sub_style = ParagraphStyle("sub",
            parent=styles["Normal"], fontSize=13, spaceAfter=20,
            textColor=HexColor("#94A3B8"), alignment=TA_CENTER)
        h2_style = ParagraphStyle("h2",
            parent=styles["Heading2"], fontSize=16, spaceBefore=14,
            spaceAfter=8, textColor=ACCENT)
        body_style = ParagraphStyle("body",
            parent=styles["Normal"], fontSize=10, spaceAfter=6,
            textColor=black, leading=16)

        # ── Cover ──
        story.append(Spacer(1, 2*cm))
        story.append(Paragraph("⚡ TrendPulse", title_style))
        story.append(Paragraph("Content Intelligence Report", sub_style))
        story.append(Spacer(1, 0.3*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=ACCENT))
        story.append(Spacer(1, 0.5*cm))

        meta_data = [
            ["Workspace",  ws_name],
            ["Niche",      niche.replace("_"," ").title()],
            ["Generated",  datetime.now().strftime("%B %d, %Y at %H:%M")],
        ]
        meta_table = Table(meta_data, colWidths=[4*cm, 12*cm])
        meta_table.setStyle(TableStyle([
            ("FONTSIZE",    (0,0), (-1,-1), 10),
            ("TEXTCOLOR",   (0,0), (0,-1), ACCENT),
            ("TEXTCOLOR",   (1,0), (1,-1), black),
            ("FONTNAME",    (0,0), (0,-1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [HexColor("#F8FAFC"), white]),
            ("GRID",        (0,0), (-1,-1), 0.3, HexColor("#E2E8F0")),
            ("PADDING",     (0,0), (-1,-1), 8),
        ]))
        story.append(meta_table)
        story.append(PageBreak())

        # ── Trends Section ──
        story.append(Paragraph("🔥 Trending Topics", h2_style))
        story.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#E2E8F0")))
        story.append(Spacer(1, 0.3*cm))

        for bucket_name, bucket_label in [
            ("exploding", "🔥 Exploding"), ("growing", "↑ Growing"),
            ("future", "◈ Future"), ("stable", "· Stable")
        ]:
            bucket = trends.get(bucket_name, [])
            if not bucket:
                continue
            story.append(Paragraph(bucket_label, ParagraphStyle("bl",
                parent=styles["Normal"], fontSize=12,
                textColor=ACCENT, fontName="Helvetica-Bold", spaceBefore=10)))
            for i, t in enumerate(bucket[:5]):
                topics = t.get("top_topics", [])
                title  = topics[0] if topics else f"Cluster #{t.get('cluster_id','')}"
                score  = t.get("cluster_score", 0)
                story.append(Paragraph(
                    f"{i+1}. {title}  <font color='#94A3B8'>(score: {score:.1f})</font>",
                    body_style
                ))

        story.append(PageBreak())

        # ── Strategy Section ──
        story.append(Paragraph("📋 Content Strategy", h2_style))
        story.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#E2E8F0")))
        story.append(Spacer(1, 0.3*cm))

        for line in (strategy or "No strategy generated yet.").split("\n"):
            line = line.strip()
            if not line:
                story.append(Spacer(1, 0.2*cm))
            elif line.startswith("##"):
                story.append(Paragraph(line.replace("#","").strip(), h2_style))
            elif line.startswith("#"):
                story.append(Paragraph(line.replace("#","").strip(),
                    ParagraphStyle("h3", parent=styles["Heading3"],
                                   fontSize=13, textColor=ACCENT)))
            else:
                story.append(Paragraph(line, body_style))

        # ── Content Ideas ──
        if ideas:
            story.append(PageBreak())
            story.append(Paragraph("✦ Generated Content Ideas", h2_style))
            story.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#E2E8F0")))
            for i, idea in enumerate(ideas):
                story.append(Spacer(1, 0.4*cm))
                story.append(Paragraph(f"Idea {idea.get('idea_index', i+1)}", ParagraphStyle("ih",
                    parent=styles["Normal"], fontSize=11,
                    textColor=ACCENT, fontName="Helvetica-Bold")))
                for key, label in [("hook","Hook"), ("post_copy","Copy"), ("hashtags","Hashtags")]:
                    val = idea.get(key, "")
                    if val:
                        if isinstance(val, list):
                            val = " ".join(f"#{t}" for t in val)
                        story.append(Paragraph(f"<b>{label}:</b> {val}", body_style))

        doc.build(story)
        logger.info(f"PDF exported: {path}")
        return path

    def _export_pdf_plain(self, ws_name, niche, strategy, trends) -> str:
        """Fallback بدون reportlab — نص عادي."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path      = os.path.join(EXPORT_DIR, f"content_plan_{timestamp}.txt")
        lines = [
            "═" * 60,
            "TrendPulse — Content Intelligence Report",
            "═" * 60,
            f"Workspace: {ws_name}",
            f"Niche:     {niche}",
            f"Date:      {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "",
            "── TRENDS ──",
        ]
        for bucket in ["exploding","growing","future"]:
            items = trends.get(bucket, [])
            if items:
                lines.append(f"\n[{bucket.upper()}]")
                for t in items[:5]:
                    topics = t.get("top_topics", [])
                    title  = topics[0] if topics else f"Cluster #{t.get('cluster_id','')}"
                    lines.append(f"  • {title}")
        lines += ["", "── STRATEGY ──", strategy or "No strategy yet."]
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        logger.info(f"Plain text export: {path}")
        return path

    # ══════════════════════════════════════════
    # Excel Export
    # ══════════════════════════════════════════

    def export_excel(self, workspace_name: str, niche: str,
                     strategy: str, trends: dict,
                     content_ideas: list = None,
                     runs: list = None) -> str:
        """
        يولّد Excel بـ 4 sheets:
        - Overview    : معلومات عامة
        - Trends      : كل الترندات مع scores
        - Content Plan: الأفكار والـ hooks
        - Run History : سجل الـ pipeline runs
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter

            return self._export_excel_openpyxl(
                workspace_name, niche, strategy, trends,
                content_ideas or [], runs or []
            )
        except ImportError:
            logger.warning("openpyxl not installed — using CSV fallback")
            return self._export_csv(workspace_name, niche, trends)

    def _export_excel_openpyxl(self, ws_name, niche, strategy, trends,
                                 ideas, runs) -> str:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path      = os.path.join(EXPORT_DIR, f"content_plan_{timestamp}.xlsx")

        wb      = openpyxl.Workbook()
        ACCENT  = "3B82F6"
        DARK    = "0D1320"
        LIGHT   = "EFF6FF"

        def header_style(cell):
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor=ACCENT)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def auto_width(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        # ── Sheet 1: Overview ──
        ws1 = wb.active
        ws1.title = "Overview"
        overview_data = [
            ["Field",       "Value"],
            ["Workspace",   ws_name],
            ["Niche",       niche.replace("_"," ").title()],
            ["Generated",   datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Exploding",   len(trends.get("exploding", []))],
            ["Growing",     len(trends.get("growing", []))],
            ["Future",      len(trends.get("future", []))],
            ["Total Runs",  len(runs)],
        ]
        for r, row in enumerate(overview_data, 1):
            for c, val in enumerate(row, 1):
                cell = ws1.cell(row=r, column=c, value=val)
                if r == 1:
                    header_style(cell)
                elif c == 1:
                    cell.font = Font(bold=True, color=ACCENT)
        auto_width(ws1)

        # ── Sheet 2: Trends ──
        ws2 = wb.create_sheet("Trends")
        headers = ["#", "Title", "State", "Forecast", "Score", "Cluster"]
        for c, h in enumerate(headers, 1):
            header_style(ws2.cell(row=1, column=c, value=h))

        row = 2
        for bucket in ["exploding","growing","future","stable"]:
            for t in trends.get(bucket, []):
                topics = t.get("top_topics", [])
                title  = topics[0] if topics else f"Cluster #{t.get('cluster_id','')}"
                ws2.cell(row=row, column=1, value=row-1)
                ws2.cell(row=row, column=2, value=title[:100])
                ws2.cell(row=row, column=3, value=t.get("cluster_state", bucket))
                ws2.cell(row=row, column=4, value=t.get("forecast", ""))
                ws2.cell(row=row, column=5, value=round(t.get("cluster_score", 0), 2))
                ws2.cell(row=row, column=6, value=t.get("cluster_id", ""))
                # color by state
                fill_colors = {
                    "exploding": "FEE2E2", "growing": "FEF9C3",
                    "future":    "EDE9FE", "stable":  "F1F5F9"
                }
                fill_color = fill_colors.get(t.get("cluster_state", bucket), "FFFFFF")
                for c in range(1, 7):
                    ws2.cell(row=row, column=c).fill = PatternFill("solid", fgColor=fill_color)
                row += 1
        auto_width(ws2)

        # ── Sheet 3: Content Plan ──
        ws3 = wb.create_sheet("Content Plan")
        plan_headers = ["#", "Hook", "Post Copy", "Hashtags", "Image Description", "Status"]
        for c, h in enumerate(plan_headers, 1):
            header_style(ws3.cell(row=1, column=c, value=h))

        for i, idea in enumerate(ideas, 1):
            tags = " ".join(f"#{t}" for t in idea.get("hashtags", []))
            ws3.cell(row=i+1, column=1, value=idea.get("idea_index", i))
            ws3.cell(row=i+1, column=2, value=idea.get("hook", "")[:100])
            ws3.cell(row=i+1, column=3, value=idea.get("post_copy", "")[:200])
            ws3.cell(row=i+1, column=4, value=tags)
            ws3.cell(row=i+1, column=5, value=idea.get("image_description", "")[:150])
            ws3.cell(row=i+1, column=6, value=idea.get("status", ""))
        auto_width(ws3)

        # ── Sheet 4: Strategy ──
        ws4 = wb.create_sheet("Strategy")
        ws4.cell(row=1, column=1, value="Content Strategy")
        header_style(ws4.cell(row=1, column=1))
        for i, line in enumerate((strategy or "").split("\n"), 2):
            ws4.cell(row=i, column=1, value=line)
        auto_width(ws4)

        wb.save(path)
        logger.info(f"Excel exported: {path}")
        return path

    def _export_csv(self, ws_name, niche, trends) -> str:
        """CSV fallback."""
        import csv
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path      = os.path.join(EXPORT_DIR, f"trends_{timestamp}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Title", "State", "Score", "Forecast"])
            for bucket in ["exploding","growing","future","stable"]:
                for t in trends.get(bucket, []):
                    topics = t.get("top_topics", [])
                    title  = topics[0] if topics else f"Cluster #{t.get('cluster_id','')}"
                    w.writerow([title, bucket, t.get("cluster_score",0), t.get("forecast","")])
        return path
